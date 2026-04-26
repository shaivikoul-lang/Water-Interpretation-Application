#!/usr/bin/env python3
"""
Read HUD ZIP–COUNTY workbook, keep Colorado rows only, join county names
from CDPHE drinking water CSV (county_fips + county), emit zip_to_county_co.json.

Usage:
  python scripts/build_zip_county_co.py
  python scripts/build_zip_county_co.py --hud Data/ZIP_COUNTY_122025.xlsx \\
      --cdphe "Data/EPHT_REF_COEPHT Public Drinking Water Data_2023_EN.xlsx - Public Drinking Water Data.csv" \\
      --out Data/zip_to_county_co.json
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def norm_zip(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        value = int(value)
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return None
    if s.endswith(".0"):
        s = s[:-2]
    if not s.isdigit():
        return None
    if len(s) > 5:
        return None
    return s.zfill(5)


def norm_county_fips(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        value = int(value)
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if not s.isdigit():
        return None
    return s.zfill(5)


def load_cdph_county_names(cdphe_path: Path) -> dict[str, str]:
    """Map 5-digit county FIPS -> CDPHE county name (mode if duplicates)."""
    pairs: list[tuple[str, str]] = []
    with cdphe_path.open(newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if (row.get("state") or "").strip() != "Colorado":
                continue
            fips = norm_county_fips(row.get("county_fips"))
            name = (row.get("county") or "").strip()
            if not fips or not name:
                continue
            if not fips.startswith("08"):
                continue
            pairs.append((fips, name))

    by_fips: dict[str, Counter] = {}
    for fips, name in pairs:
        by_fips.setdefault(fips, Counter())[name] += 1

    return {fips: c.most_common(1)[0][0] for fips, c in by_fips.items()}


def load_hud_zip_county(
    hud_path: Path,
    fips_to_name: dict[str, str],
) -> dict[str, list[dict]]:
    """ZIP (str) -> list of county records, CO only, sorted by tot_ratio desc."""
    wb = load_workbook(hud_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise SystemExit("HUD workbook is empty")

    col_index = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    for required in ("ZIP", "COUNTY", "RES_RATIO", "TOT_RATIO"):
        if required not in col_index:
            raise SystemExit(f"Missing column {required!r} in HUD sheet. Found: {list(col_index)}")

    i_zip = col_index["ZIP"]
    i_county = col_index["COUNTY"]
    i_res = col_index["RES_RATIO"]
    i_tot = col_index["TOT_RATIO"]

    out: dict[str, list[dict]] = {}

    for tup in rows:
        if not tup or tup[i_zip] is None:
            continue
        z = norm_zip(tup[i_zip])
        cf = norm_county_fips(tup[i_county])
        if not z or not cf or not cf.startswith("08"):
            continue

        def to_float(x):
            if x is None or x == "":
                return None
            try:
                return float(x)
            except (TypeError, ValueError):
                return None

        res = to_float(tup[i_res])
        tot = to_float(tup[i_tot])
        name = fips_to_name.get(cf)
        rec = {
            "county_fips": cf,
            "county_name": name,
            "res_ratio": res,
            "tot_ratio": tot,
        }
        out.setdefault(z, []).append(rec)

    wb.close()

    for z, lst in out.items():
        lst.sort(key=lambda r: (r["tot_ratio"] is not None, r["tot_ratio"] or 0), reverse=True)

    return out


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    default_hud = root / "Data" / "ZIP_COUNTY_122025.xlsx"
    default_cdphe = (
        root
        / "Data"
        / "EPHT_REF_COEPHT Public Drinking Water Data_2023_EN.xlsx - Public Drinking Water Data.csv"
    )
    default_out = root / "Data" / "zip_to_county_co.json"

    ap = argparse.ArgumentParser(description="Build Colorado ZIP→county JSON from HUD + CDPHE names.")
    ap.add_argument("--hud", type=Path, default=default_hud, help="HUD ZIP-COUNTY .xlsx")
    ap.add_argument("--cdphe", type=Path, default=default_cdphe, help="CDPHE CSV for county name join")
    ap.add_argument("--out", type=Path, default=default_out, help="Output JSON path")
    args = ap.parse_args()

    if not args.hud.is_file():
        raise SystemExit(f"HUD file not found: {args.hud}")
    if not args.cdphe.is_file():
        raise SystemExit(f"CDPHE file not found: {args.cdphe}")

    fips_to_name = load_cdph_county_names(args.cdphe)
    zips = load_hud_zip_county(args.hud, fips_to_name)

    missing_names = sum(
        1 for z, lst in zips.items() for r in lst if r["county_name"] is None
    )

    payload = {
        "source": "HUD USPS ZIP–COUNTY crosswalk (filtered to Colorado)",
        "source_file": args.hud.name,
        "county_name_join": "CDPHE Public Drinking Water Data (unique county_fips → county)",
        "cdphe_file": args.cdphe.name,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "zip_count": len(zips),
        "county_fips_names_resolved": len(fips_to_name),
        "hud_rows_missing_cdph_county_name": missing_names,
        "zips": zips,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print(f"Wrote {args.out}")
    print(f"  Colorado ZIPs with ≥1 county row: {len(zips)}")
    print(f"  CDPHE county FIPS→name pairs: {len(fips_to_name)}")
    print(f"  HUD county rows without CDPHE name: {missing_names}")


if __name__ == "__main__":
    main()
